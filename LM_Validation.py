import psycopg2
import pandas as pd
from neo4j import GraphDatabase

from config import load_config
from logger import get_logger
from Queries import (
    loc, loc_per_tech, dlms,
    extension_count, missing_code_db,
    analyzed_files, critical_violations,
    missing_code, check_schemas,
    loc_null, customized_jobs,
    fetch_app_schema
)

# ------------------ LOGGER ------------------
logger = get_logger(__name__)

# ------------------ LOAD CONFIG ------------------
logger.info("Loading configuration")
config = load_config()

# ------------------ POSTGRES CONNECTION ------------------
def postgres_connection():
    try:
        logger.info("Connecting to PostgreSQL")
        return psycopg2.connect(
            host=config['CSS_HOST'],
            port=config['CSS_PORT'],
            dbname=config['CSS_DB'],
            user=config['CSS_USERNAME'],
            password=config['CSS_PASSWORD']
        )
    except Exception:
        logger.exception("PostgreSQL connection failed")
        raise

# ------------------ NEO4J CONNECTION ------------------
def neo4j_connection(uri, username, password):
    try:
        logger.info("Connecting to Neo4j")
        return GraphDatabase.driver(uri, auth=(username, password))
    except Exception:
        logger.exception("Neo4j connection failed")
        raise

# ------------------ FETCH NEO4J OBJECT COUNTS ------------------
def fetch_neo4j_object_counts(driver, database_names):
    logger.info("Fetching Neo4j object counts")
    app_object_counts = {}

    for db in database_names:
        logger.info(f"[Neo4j DB={db}] Processing")

        with driver.session(database=db) as session:
            apps = session.run("""
                MATCH(n:Application)
RETURN n.DisplayName as consoleApp_name ,n.Name as app_name
            """)

            for record in apps:
                app_name = record["app_name"]
                capp_name=record["consoleApp_name"]
                if not app_name:
                    continue

                query = f"""
                MATCH (o:Object:`{app_name}`)
                WHERE NOT 'Deleted' IN labels(o)
                RETURN count(o) AS cnt
                """

                result = session.run(query).single()
                count = result["cnt"] if result else 0
                app_object_counts[capp_name] = (
                    app_object_counts.get(capp_name, 0) + count
                )

    logger.info("Neo4j object count collection completed")
    return app_object_counts

# ------------------ VALUE EXTRACTION ------------------
def extract_v2(param, value):
    if not value:
        return 0

    if param == "loc":
        return value[0][0]

    if param == "loc_per_tech":
        return ", ".join(f"{tech}:{int(loc)}" for tech, loc in value)

    if param in (
        "extension_count", "dlms",
        "analyzed_files", "Total Object Count",
        "Customized Jobs"
    ):
        return value[0][0]

    if param in ("missing_code_db", "Missing Code"):
        return value[0][0]

    if param == "Dashboard - Critical violations":
        return value[0][2]

    return 0

# ------------------ BUILD EXCEL ROWS ------------------
def build_excel_rows(all_data):
    rows = []
    app_name = all_data["app_name"]
    first = True

    for key, value in all_data.items():
        if key in ("domain_name", "app_name"):
            continue

        rows.append({
            "App Name": app_name if first else "",
            "Parameters": key.replace("_", " ").title(),
            "V2": extract_v2(key, value),
            "V3": "",
            "Variation": ""
        })
        first = False

    return rows
def build_excel_rows_v3(all_data):
    rows = []
    app_name = all_data["app_name"]
    first = True

    for key, value in all_data.items():
        if key in ("domain_name", "app_name"):
            continue

        rows.append({
            "App Name": app_name if first else "",
            "Parameters": key.replace("_", " ").title(),
            "V2":"",
            "V3":extract_v2(key, value),
            "Variation": ""
        })
        first = False

    return rows
# ------------------ MAIN REPORT ------------------
def generate_report():
    logger.info("V3 Upgrade Validation started")

    connection = postgres_connection()
    cursor = connection.cursor()

    neo4j_driver = neo4j_connection(
        config['NEO4J_URL'],
        config['NEO4J_USER'],
        config['NEO4J_PASSWORD']
    )

    tenants = config["NEO4J_DB"].split(',')
    neo4j_object_counts = fetch_neo4j_object_counts(
        neo4j_driver, tenants
    )
    cursor.execute(
        "SELECT guid, name FROM aip_node.domain ORDER BY guid ASC"
    )
    domains = cursor.fetchall()

    with pd.ExcelWriter(
        "V3_Upgrade_Apps_Validation.xlsx",
        engine="openpyxl"
    ) as writer:

        # ---- TRACK DEFAULT APPS TO SKIP DUPLICATES ----
        processed_default_apps = set()

        for domain_guid, domain_name in domains:
            logger.info(
                f"[Domain={domain_name} | GUID={domain_guid}] Starting domain processing"
            )

            cursor.execute(fetch_app_schema, (domain_guid,))
            apps = cursor.fetchall()

            cursor.execute(check_schemas)

            for app_name, schema, app_domain_guid in apps:
                sheet = domain_name if app_domain_guid else "default"

                # ---- SKIP DUPLICATE DEFAULT APPS ----
                if sheet == "default":
                    if app_name in processed_default_apps:
                        logger.warning(
                            f"[Domain=default | App={app_name}] Skipping duplicate application"
                        )
                        continue
                    processed_default_apps.add(app_name)

                context = f"[Domain={sheet} | App={app_name} | Schema={schema}]"
                logger.info(f"{context} Starting application processing")

                try:
                    # -------- CENTRAL --------
                    cursor.execute(f"SET search_path TO {schema}_central")

                    loc_query = loc_null if app_domain_guid is None else loc
                    if app_domain_guid is None:
                        cursor.execute(loc_query, (app_name,))
                    else:
                        cursor.execute(loc_query, (app_domain_guid, app_name))

                    loc_rows = cursor.fetchall()
                    logger.info(f" LOC rows for application {app_name}: {loc_rows}")

                    cursor.execute(loc_per_tech)
                    loc_per_tech_rows = cursor.fetchall()
                    logger.info(f" LOC per tech rows for application {app_name}: {loc_per_tech_rows}")

                    cursor.execute(extension_count)
                    extension_count_rows = cursor.fetchall()
                    logger.info(f" Extension rows for application {app_name}: {extension_count_rows}")

                    cursor.execute(critical_violations)
                    critical_violations_rows = cursor.fetchall()
                    logger.info(f" critical violations rows for application {app_name}: {critical_violations_rows}")

                    # -------- LOCAL --------
                    cursor.execute(f"SET search_path TO {schema}_local")
                    cursor.execute(dlms)
                    dlms_rows = cursor.fetchall()
                    logger.info(f" DLMS rows: for application {app_name} {dlms_rows}")
                    cursor.execute(missing_code_db)
                    missing_code_db_rows = cursor.fetchall()
                    logger.info(f" Missing Code DB rowsfor application {app_name}: {len(missing_code_db_rows)}")
                    cursor.execute(analyzed_files)
                    analyzed_files_rows = cursor.fetchall()
                    logger.info(f" Analyzed Files rows for application {app_name}: {analyzed_files_rows}")
                    cursor.execute(missing_code)
                    missing_codes = cursor.fetchall()
                    logger.info(f" Missing Codes for application {app_name}: {len(missing_codes)}")

                    # -------- MNGT --------
                    cursor.execute(f"SET search_path TO {schema}_mngt")
                    cursor.execute(customized_jobs)
                    customized_jobs_rows = cursor.fetchall()
                    logger.info("customized_jobs_rows for for application {} : {}".format(app_name,customized_jobs_rows))
                    total_object_count = neo4j_object_counts.get(app_name, 0)

                    if app_name in neo4j_object_counts:
                        logger.info(f"Total objects for application '{app_name}' found: {total_object_count} in Neoej applications {neo4j_object_counts}" )
                    else:
                        logger.warning(
                            f" Total objects for Application '{app_name}' not found in Neo4j object counts. Defaulting to 0")

                except Exception:
                    logger.exception(f"{context} ERROR during processing")
                    continue

                logger.info(
                    f"{context} Completed successfully | Neo4j Objects={total_object_count}"
                )

                all_data = {
                    "domain_name": sheet,
                    "app_name": app_name,
                    "loc": loc_rows,
                    "loc_per_tech": loc_per_tech_rows,
                    "extension_count": extension_count_rows,
                    "dlms": dlms_rows,
                    "missing_code_db": missing_code_db_rows,
                    "analyzed_files": analyzed_files_rows,
                    "Missing Code": missing_codes,
                    "Dashboard - Critical violations": critical_violations_rows,
                    "Total Object Count": [(total_object_count,)],
                    "Customized Jobs": customized_jobs_rows
                }

                df = pd.DataFrame(build_excel_rows(all_data))
                print(df)
                startrow = (
                    writer.sheets[sheet].max_row
                    if sheet in writer.sheets else 0
                )

                df.to_excel(
                    writer,
                    sheet_name=sheet,
                    index=False,
                    startrow=startrow,
                    header=startrow == 0
                )

    cursor.close()
    connection.close()
    neo4j_driver.close()

    logger.info("V3_Upgrade_Apps_Validation.xlsx generated successfully")

def generate_report3():
    logger.info("V3 Upgrade Validation started")

    excel_file = "V3_Upgrade_Apps_Validation.xlsx"

    connection = postgres_connection()
    cursor = connection.cursor()

    neo4j_driver = neo4j_connection(
        config['V3_NEO4J_URL'],
        config['V3_NEO4J_USER'],
        config['V3_NEO4J_PASSWORD']
    )

    tenants = config["V3_NEO4J_DB"].split(',')
    neo4j_object_counts = fetch_neo4j_object_counts(
        neo4j_driver, tenants
    )

    cursor.execute(
        "SELECT guid, name FROM aip_node.domain ORDER BY guid ASC"
    )
    domains = cursor.fetchall()
    processed_default_apps = set()

    for domain_guid, domain_name in domains:
        logger.info(
            f"[Domain={domain_name} | GUID={domain_guid}] Starting domain processing"
        )

        cursor.execute(fetch_app_schema, (domain_guid,))
        apps = cursor.fetchall()

        cursor.execute(check_schemas)

        for app_name, schema, app_domain_guid in apps:
            sheet = domain_name if app_domain_guid else "default"

            # ---- SKIP DUPLICATE DEFAULT APPS ----
            if sheet == "default":
                if app_name in processed_default_apps:
                    logger.warning(
                        f"[Domain=default | App={app_name}] Skipping duplicate application"
                    )
                    continue
                processed_default_apps.add(app_name)

            context = f"[Domain={sheet} | App={app_name} | Schema={schema}]"
            logger.info(f"{context} Starting application processing")

            try:
                # -------- CENTRAL --------
                cursor.execute(f"SET search_path TO {schema}_central")

                loc_query = loc_null if app_domain_guid is None else loc
                if app_domain_guid is None:
                    cursor.execute(loc_query, (app_name,))
                else:
                    cursor.execute(loc_query, (app_domain_guid, app_name))
                loc_rows = cursor.fetchall()

                cursor.execute(loc_per_tech)
                loc_per_tech_rows = cursor.fetchall()

                cursor.execute(extension_count)
                extension_count_rows = cursor.fetchall()

                cursor.execute(critical_violations)
                critical_violations_rows = cursor.fetchall()

                # -------- LOCAL --------
                cursor.execute(f"SET search_path TO {schema}_local")
                cursor.execute(dlms)
                dlms_rows = cursor.fetchall()

                cursor.execute(missing_code_db)
                missing_code_db_rows = cursor.fetchall()

                cursor.execute(analyzed_files)
                analyzed_files_rows = cursor.fetchall()

                cursor.execute(missing_code)
                missing_codes = cursor.fetchall()

                # -------- MNGT --------
                cursor.execute(f"SET search_path TO {schema}_mngt")
                cursor.execute(customized_jobs)
                customized_jobs_rows = cursor.fetchall()

                total_object_count = neo4j_object_counts.get(app_name, 0)

            except Exception:
                logger.exception(f"{context} ERROR during processing")
                continue

            # ---- build all_data (UNCHANGED STRUCTURE) ----
            all_data = {
                "domain_name": sheet,
                "app_name": app_name,
                "loc": loc_rows,
                "loc_per_tech": loc_per_tech_rows,
                "extension_count": extension_count_rows,
                "dlms": dlms_rows,
                "missing_code_db": missing_code_db_rows,
                "analyzed_files": analyzed_files_rows,
                "Missing Code": missing_codes,
                "Dashboard - Critical violations": critical_violations_rows,
                "Total Object Count": [(total_object_count,)],
                "Customized Jobs": customized_jobs_rows
            }
            # Read existing sheet
            # Read existing sheet
            df = pd.read_excel(excel_file, sheet_name=sheet)

            # Fill empty App Name cells so all rows of an app have the app name
            df["App Name"] = df["App Name"].replace('', pd.NA).ffill()

            # Ensure V3 column can accept strings
            df["V3"] = df["V3"].astype(object)

            # Collect all V3 data for this sheet
            excel_rows = pd.DataFrame(build_excel_rows_v3(all_data))
            # Ensure App Name is filled for all rows
            excel_rows["App Name"] = excel_rows["App Name"].replace('', pd.NA).ffill()

            # Iterate over all unique apps in V3 data
            for app in excel_rows["App Name"].unique():
                mask = df["App Name"] == app
                if mask.any():
                    # Get V3 values for this app from excel_rows
                    v3_values = excel_rows[excel_rows["App Name"] == app]["V3"].tolist()

                    # Make sure lengths match; if fewer values, repeat last; if more, truncate
                    if len(v3_values) < mask.sum():
                        v3_values += [v3_values[-1]] * (mask.sum() - len(v3_values))
                    elif len(v3_values) > mask.sum():
                        v3_values = v3_values[:mask.sum()]

                    # Assign row by row to the V3 column in Excel
                    df.loc[mask, "V3"] = v3_values
                else:
                    logger.warning(f"App '{app}' not found in sheet '{sheet}'")

            # Write back updated sheet
            with pd.ExcelWriter(
                    excel_file,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)

        # ---- write back same sheet ----
        # with pd.ExcelWriter(
        #     excel_file,
        #     engine="openpyxl",
        #     mode="a",
        #     if_sheet_exists="replace"
        # ) as writer:
        #     df.to_excel(writer, sheet_name=sheet, index=False)

    cursor.close()
    connection.close()
    neo4j_driver.close()

    logger.info("V3 Upgrade Validation completed using build_excel_rows")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def calculate_variation_only_clean(excel_file):
    """
    Update only the 'Variation' column in all sheets,
    calculated as V2 - V3, supports numeric and tech breakdown strings like 'JEE:59803, SQL:0'.
    """
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_font = Font(bold=True)

    def calculate_string_variation(v2_str, v3_str):
        """Calculate variation for strings like 'JEE:59803, SQL:0'."""
        if not v2_str or not v3_str:
            return ""
        try:
            parts_v2 = [p.strip() for p in str(v2_str).split(",")]
            parts_v3 = [p.strip() for p in str(v3_str).split(",")]
            dict_v2 = {p.split(":")[0]: float(p.split(":")[1]) for p in parts_v2 if ":" in p}
            dict_v3 = {p.split(":")[0]: float(p.split(":")[1]) for p in parts_v3 if ":" in p}
            variation_dict = {k: dict_v2.get(k, 0) - dict_v3.get(k, 0) for k in dict_v2.keys()}
            variation_str = ", ".join(f"{k}:{int(v)}" for k, v in variation_dict.items())
            return variation_str
        except:
            return ""

    wb = load_workbook(excel_file)

    for sheet in wb.sheetnames:
        print(f"Processing sheet: {sheet}")
        ws = wb[sheet]

        # Find column indexes
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}  # 1-based
        col_v2 = headers.get("V2")
        col_v3 = headers.get("V3")
        col_variation = headers.get("Variation")

        if not col_v2 or not col_v3:
            print(f"Skipping sheet '{sheet}': V2 or V3 column not found")
            continue

        # If Variation column does not exist, create it at the end
        if not col_variation:
            col_variation = ws.max_column + 1
            ws.cell(row=1, column=col_variation, value="Variation")

        # Calculate Variation for each row
        for row in range(2, ws.max_row + 1):
            v2 = ws.cell(row=row, column=col_v2).value
            v3 = ws.cell(row=row, column=col_v3).value

            # Check if value is a tech breakdown string
            if isinstance(v2, str) and ":" in v2:
                variation = calculate_string_variation(v2, v3)
            else:
                try:
                    v2_num = float(v2) if v2 is not None else 0
                    v3_num = float(v3) if v3 is not None else 0

                    # If both are zero → variation = 0
                    if v2_num == 0 and v3_num == 0:
                        variation = 0
                    elif v2_num == 0:
                        # Avoid division by zero → just put v3 value or leave empty
                        variation = 0
                    else:
                        # Normal percentage variation
                        variation = round(((v3_num - v2_num) / v2_num) * 100, 2)
                except:
                    variation = ""  # leave empty if not numeric

            ws.cell(row=row, column=col_variation, value=variation)


            # Conditional formatting: red + bold if numeric variation exceeds ±5
            try:
                if isinstance(variation, (int, float)) and abs(variation) > 5:
                    cell = ws.cell(row=row, column=col_variation)
                    cell.fill = red_fill
                    cell.font = bold_font
            except:
                continue

    wb.save(excel_file)
    print("Variation column updated for all sheets, including tech breakdowns.")


def main_menu():
    while True:
        print("\nPlease choose an option:")
        print("1: Generate V2 report")
        print("2: Generate V3 report")
        print("3: Calculate Variation")
        print("0: Exit")
        try:
            choice = int(input("Enter your choice: "))
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue

        if choice == 1:
            generate_report()  # Your existing V2 function
        elif choice == 2:
            generate_report3()  # Your existing V3 function
        elif choice == 3:
            calculate_variation_only_clean("V3_Upgrade_Apps_Validation.xlsx")
        elif choice == 0:
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please enter 0, 1, 2, or 3.")

        # Ask if user wants to continue
        cont = input("Do you want to continue? (Y/N): ").strip().lower()
        if cont != 'y':
            print("Exiting...")
            break


if __name__ == "__main__":
    main_menu()


